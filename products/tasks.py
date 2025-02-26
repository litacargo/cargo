import os
import openpyxl
from celery import shared_task
from datetime import datetime, timedelta
from django.utils import timezone
from .models import Product, Status, Client, ProductFile
from config.models import Configuration
from config.config import BaseStatus

from notifications.notification_transit import notification_for_products_status_transit

from notifications.notification_china import send_notification_china
from notifications.notification_bishkek import send_notification_bihskek

from openpyxl import load_workbook

import logging
logger = logging.getLogger(__name__)




@shared_task
def update_product_statuses():
    transit_hours = Configuration.objects.get(key="transit_hours")
    hours_value = int(transit_hours.value)
    hours_ago = timezone.now() - timedelta(hours=hours_value) # Получаем дату, которая находится на transit_hours часов назад
    status_china = Status.objects.get(name=BaseStatus.CHINA)
    products_in_china = Product.objects.filter(status=status_china, date__lte=hours_ago) # Получаем товары в Китае, которые находятся там более 2 дней

    updated_count = 0 # Количество обновленных товаров
    updated_ids = [] # Список обновленных товаров

    for product in products_in_china:
        product.status = Status.objects.get(name='В пути') # Обновляем статус
        product.date = timezone.now()  # Обновляем дату на текущую
        product.save() # Сохраняем изменения
        updated_count += 1  # Увеличиваем счетчик
        updated_ids.append(product.id) # Добавляем ID товара в список

    
    notification_for_products_status_transit.delay()
    return {
        'updated_count': updated_count,
        'updated_ids': updated_ids,
        'timestamp': datetime.now().isoformat(),
    }



@shared_task
def process_china_products(product_file_id, request_user_id):
    """
    Асинхронная обработка файла с товарами для Китая.
    """
    try:
        # Получаем объект файла из базы данных
        product_file = ProductFile.objects.get(id=product_file_id)
        file_path = product_file.file.path
        logger.info(f"Обрабатываем файл: {file_path}")

        if not os.path.exists(file_path):
            logger.error(f"Файл не найден: {file_path}")
            product_file.status = 'failed'
            product_file.error_message = f"Файл не найден: {file_path}"
            product_file.save()
            return {'error': f"Нет такого файла: '{file_path}'"}

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        products_created = 0
        products_skipped = 0
        clients_products_count = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 1 or not row[0]:
                continue

            product_code = str(row[0]).strip()
            client_id = None
            client = None

            if len(row) >= 2 and row[1] is not None:
                client_id = row[1]
                if isinstance(client_id, float):
                    client_id = str(int(client_id))
                else:
                    client_id = str(client_id).strip()

                client = Client.objects.filter(code=client_id).first()

            china_status, _ = Status.objects.get_or_create(name=BaseStatus.CHINA)

            if client and Product.objects.filter(product_code=product_code, client=client).exists():
                products_skipped += 1
                continue
            elif not client and Product.objects.filter(product_code=product_code, client__isnull=True).exists():
                products_skipped += 1
                continue

            Product.objects.create(
                product_code=product_code,
                client=client,
                date=datetime.utcnow(),
                status=china_status,
                branch=client.branch if client else None
            )
            products_created += 1

            if client and client.id not in clients_products_count:
                clients_products_count[client.id] = 0
            if client:
                clients_products_count[client.id] += 1

        if clients_products_count:
            send_notification_china.delay(clients_products_count)

        # Обновляем статус файла
        product_file.status = 'processed'
        product_file.save()

        result = {
            "products_created": products_created,
            "products_skipped": products_skipped,
            "clients_products_count": clients_products_count
        }
        logger.info(f"Обработка завершена: {result}")
        return result

    except ProductFile.DoesNotExist:
        logger.error(f"Файл с ID {product_file_id} не найден в базе данных")
        return {"error": f"Файл с ID {product_file_id} не найден"}
    except Exception as e:
        logger.error(f"Ошибка при обработке файла ID {product_file_id}: {e}")
        product_file.status = 'failed'
        product_file.error_message = str(e)
        product_file.save()
        return {"error": str(e)}
    finally:
        # Удаляем временный файл после обработки
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"Файл удалён: {file_path}")

@shared_task
def process_bishkek_products(product_file_id, request_user_id):
    """
    Асинхронная обработка файла с товарами для Бишкека.
    """
    try:
        # Получаем объект файла из базы данных
        product_file = ProductFile.objects.get(id=product_file_id)
        file_path = product_file.file.path
        logger.info(f"Обрабатываем файл: {file_path}")

        if not os.path.exists(file_path):
            logger.error(f"Файл не найден: {file_path}")
            product_file.status = 'failed'
            product_file.error_message = f"Файл не найден: {file_path}"
            product_file.save()
            return {'error': f"Нет такого файла: '{file_path}'"}

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        products_created = 0
        products_updated = 0
        clients_products_count = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 2 or not row[0] or not row[2]:
                continue

            product_code = str(row[0]).strip()
            client_id = row[1] if len(row) > 1 else None
            weight = row[2]
            price = row[3] if len(row) > 3 else None

            client = None
            if client_id:
                if isinstance(client_id, float):
                    client_id = str(int(client_id))
                else:
                    client_id = str(client_id).strip()

                client = Client.objects.filter(code=client_id).first()

            bishkek_status, _ = Status.objects.get_or_create(name=BaseStatus.BISHKEK)

            product, created = Product.objects.get_or_create(
                product_code=product_code,
                client=client,
                defaults={
                    "status": bishkek_status,
                    "weight": weight,
                    "date": timezone.now().date(),
                }
            )

            if created:
                products_created += 1
            else:
                update_fields = {
                    "weight": weight,
                    "status": bishkek_status,
                    "date": timezone.now().date(),
                }
                if price is not None:
                    update_fields["price"] = price

                Product.objects.filter(id=product.id).update(**update_fields)
                products_updated += 1

            if client and client.id not in clients_products_count:
                clients_products_count[client.id] = 0
            if client:
                clients_products_count[client.id] += 1

        if clients_products_count:
            send_notification_bihskek.delay(clients_products_count)

        # Обновляем статус файла
        product_file.status = 'processed'
        product_file.save()

        result = {
            "products_created": products_created,
            "products_updated": products_updated,
            "clients_products_count": clients_products_count
        }
        logger.info(f"Обработка завершена: {result}")
        return result

    except ProductFile.DoesNotExist:
        logger.error(f"Файл с ID {product_file_id} не найден в базе данных")
        return {"error": f"Файл с ID {product_file_id} не найден"}
    except Exception as e:
        logger.error(f"Ошибка при обработке файла ID {product_file_id}: {e}")
        product_file.status = 'failed'
        product_file.error_message = str(e)
        product_file.save()
        return {"error": str(e)}
    finally:
        # Удаляем временный файл после обработки
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"Файл удалён: {file_path}")