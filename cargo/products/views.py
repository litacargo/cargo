from datetime import datetime
from django.utils import timezone
from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required, permission_required
from django.core.paginator import Paginator
from products.models import Product, Status
from clients.models import Client
from branch.models import Branch

from django.views import View
from django.contrib import messages

from config.views import get_user_branch
from config.config import BaseStatus

import openpyxl

from notifications.notification_china import send_notification_china
from notifications.notification_bishkek import send_notification_bihskek

class ProductListView(View):
    def get(self, request):
        # Получение параметров фильтрации
        search_query = request.GET.get('search', '')
        status_filter = request.GET.get('status', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')

        # Получение всех товаров
        user = request.user
        products_list = get_user_branch(user, Product).exclude(status__name=BaseStatus.PIKED).order_by('-id')

        # Фильтрация по поисковому запросу
        if search_query:
            products_list = products_list.filter(
                product_code__icontains=search_query
            ) | products_list.filter(
                status__name__icontains=search_query
            ) | products_list.filter(
                client__name__icontains=search_query
            )

        # Фильтрация по статусу
        if status_filter:
            products_list = products_list.filter(status__id=status_filter)

        # Фильтрация по дате
        if start_date:
            products_list = products_list.filter(date__gte=start_date)
        if end_date:
            products_list = products_list.filter(date__lte=end_date)

        # Пагинация
        paginator = Paginator(products_list, 30)  # 30 товаров на страницу
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Получение всех статусов для фильтрации
        statuses = Status.objects.all()

        # Рендер страницы
        return render(
            request,
            'products/products.html',
            {
                'current_page': 'products',
                'title': 'Товары',
                'page_obj': page_obj,
                'search_query': search_query,
                'status_filter': status_filter,
                'start_date': start_date,
                'end_date': end_date,
                'statuses': statuses,
            }
        )

    
    def edit_products_page(request):
        product_ids = request.POST.getlist('product_ids')
        products = Product.objects.filter(id__in=product_ids)
        statuses = Status.objects.all()
        return render(request, 'products/edit_products.html', {
                'products': products,
                'statuses': statuses,
                'current_page': 'products',
                'title': 'Изменить',
            })
    
    def update_products(request):
        product_ids = request.POST.getlist('product_ids')
        new_status_id = request.POST.get('status', '').strip()

        if not product_ids:
            messages.error(request, "Не выбраны товары.")
            return redirect('products')

        update_data = {"date": datetime.utcnow()}

        try:
            # Если указан новый статус
            if new_status_id:
                new_status = Status.objects.get(id=new_status_id)
                update_data["status"] = new_status

            # Проверяем, есть ли изменения
            if len(update_data) > 1:  # кроме даты должно быть хотя бы одно поле
                Product.objects.filter(id__in=product_ids).update(**update_data)
                messages.success(request, "Товары успешно обновлены.")
            else:
                messages.warning(request, "Ничего не изменено, не были указаны ни новый статус, ни филиал.")

        except Status.DoesNotExist:
            messages.error(request, "Выбранный статус не существует.")
        except Branch.DoesNotExist:
            messages.error(request, "Выбранный филиал не существует.")
        except ValueError:
            messages.error(request, "Некорректные данные для обновления товаров.")
        except Exception as e:
            messages.error(request, f"Ошибка при обновлении товаров: {e}")

        return redirect('products')
    
    def delete_products(request):
        """
        Удаление выбранных товаров с подтверждением.
        """
        if request.method == "POST" and 'confirm' in request.POST:
            # Если пользователь подтвердил удаление
            product_ids = request.POST.getlist('product_ids')
            deleted_count, _ = Product.objects.filter(id__in=product_ids).delete()
            messages.success(request, f"Удалено {deleted_count} товаров.")
            return redirect('products')  # Замените 'products' на имя URL для списка товаров

        elif request.method == "POST":
            # Первоначальный запрос на удаление — отобразить страницу подтверждения
            product_ids = request.POST.getlist('product_ids')
            if not product_ids:
                messages.error(request, "Не выбраны товары для удаления.")
                return redirect('products')

            products = Product.objects.filter(id__in=product_ids)
            return render(request, 'products/confirm_delete_products.html', {
                'products': products,
                'current_page': 'products',
                'title': 'Подтверждение удаления',
            })





@login_required
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    statuses = Status.objects.all()


    if request.method == "POST":
        # Получаем данные из POST и заменяем запятую на точку
        product_code = request.POST.get("product_code", "").strip()
        weight_str = request.POST.get("weight", "").replace(",", ".").strip()
        price_str = request.POST.get("price", "").replace(",", ".").strip()
        date_str = request.POST.get("date", "").strip()
        status_id = request.POST.get("status", "").strip()
        client_code = request.POST.get("client_code", "").strip()  # Изменено на client_code
        branch_id = request.POST.get("branch", "").strip()

        # Если поле не пустое, только тогда меняем:
        if product_code:
            product.product_code = product_code
        if weight_str:
            try:
                product.weight = Decimal(weight_str)
            except (ValueError, Decimal.InvalidOperation):
                messages.error(request, "Неверный формат веса.")
                return render(request, 'products/edit_product.html', {
                    'product': product,
                    'statuses': statuses,
                    'current_page': 'products',
                    'title': 'Изменить',
                })
        if price_str:
            try:
                product.price = int(Decimal(price_str))  # IntegerField требует целое число
            except (ValueError, Decimal.InvalidOperation):
                messages.error(request, "Неверный формат цены.")
                return render(request, 'products/edit_product.html', {
                    'product': product,
                    'statuses': statuses,
                    'current_page': 'products',
                    'title': 'Изменить',
                })
        if date_str:
            product.date = date_str
        if status_id:
            product.status_id = status_id
        if client_code:
            try:
                client = Client.objects.get(code=client_code)
                product.client = client
            except Client.DoesNotExist:
                messages.error(request, "Клиент с таким кодом не найден.")
                return render(request, 'products/edit_product.html', {
                    'product': product,
                    'statuses': statuses,
                    'current_page': 'products',
                    'title': 'Изменить',
                })
        if branch_id:
            product.branch_id = branch_id

        # Сохраняем объект
        product.save()
        messages.success(request, "Ваши изменения успешно сохранены!")
        return redirect('products')

    return render(request, 'products/edit_product.html', {
        'product': product,
        'statuses': statuses,
        'current_page': 'products',
        'title': 'Изменить',
    })


@login_required
@permission_required('products.delete_product', raise_exception=True)
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == "POST":
        product.delete()
        return redirect('products')  # Замените 'products' на имя URL для списка товаров

    return render(request, 'products/delete_product.html', {'product': product, 'current_page': 'products', 'title': 'Удалить',})

@login_required
def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)  # Извлекаем клиента по id
    client = product.client

    return render(request, 'products/product_detail.html', {
        'product': product,
        'client': client,
        'current_page': 'products',
        'title': 'Детали',
    })


def page_china(request):
    return render(request, "products/china.html", {"title": "Китай", 'current_page': 'products'})

def add_products_china(request):
    """Обработка загрузки файла и добавления товаров"""
    if request.method == "POST":
        try:
            file = request.FILES.get('file')  # Получаем файл
            if not file:
                messages.error(request, "Файл не выбран.")
                return redirect("page_china")  # Перенаправляем на форму

            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            products_created = 0
            products_skipped = 0  # Счётчик пропущенных товаров
            clients_products_count = {}

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
                if len(row) < 2 or not row[0] or not row[1]:
                    continue  # Пропустить строки с недостаточными данными

                product_code = str(row[0]).strip()
                client_id = row[1]

                # Если client_id — число с плавающей точкой, убираем ".0"
                if isinstance(client_id, float):
                    client_id = str(int(client_id))
                else:
                    client_id = str(client_id).strip()

                # Найти клиента по code
                client = Client.objects.filter(code=client_id).first()
                if not client:
                    messages.warning(request, f"Нету такого клиента: {client_id}")
                    continue

                # Получить статус "Китай"
                china_status, _ = Status.objects.get_or_create(name=BaseStatus.CHINA)

                # Проверить, существует ли уже товар с этим кодом
                if Product.objects.filter(product_code=product_code, client=client).exists():
                    products_skipped += 1
                    continue

                # Создать продукт
                Product.objects.create(
                    product_code=product_code,
                    client=client,
                    date=datetime.utcnow(),
                    status=china_status,
                    branch=client.branch
                )
                products_created += 1

                # Увеличиваем счетчик товаров у клиента
                if client.id not in clients_products_count:
                    clients_products_count[client.id] = 0
                clients_products_count[client.id] += 1

            if products_created > 0:
                messages.success(request, f"Успешно создано продуктов: {products_created}")
            if products_skipped > 0:
                messages.warning(request, f"Пропущено продуктов (уже существуют): {products_skipped}")
            # Добавляем сообщение с количеством товаров по каждому клиенту
            if clients_products_count:
                send_notification_china.delay(clients_products_count)
                for client_id, count in clients_products_count.items():
                    messages.info(request, f"Клиент {client_id}: {count} товаров")

        except Exception as e:
            messages.error(request, f"Ошибка при обработке файла: {str(e)}")

        return redirect("page_china")  # Перенаправляем на форму после обработки

    return redirect("page_china")



def page_bishkek(request):
    """Отображение страницы для загрузки товаров в Бишкеке"""
    return render(request, "products/bishkek.html", {"title": "Бишкек", 'current_page': 'products'})


def add_products_bishkek(request):
    """Обработка загрузки файла и добавления/обновления товаров"""
    if request.method == "POST":  # Используем POST для загрузки
        try:
            file = request.FILES.get('file')  # Получаем файл
            if not file:
                messages.error(request, "Файл не выбран.")
                return redirect("page_bishkek")  # Перенаправляем на форму

            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            products_created = 0
            products_updated = 0  # Счётчик обновленных товаров
            clients_products_count = {}

            for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
                if len(row) < 3 or not row[0] or not row[1] or not row[2]:
                    continue  # Пропустить строки с недостаточными данными

                product_code = str(row[0]).strip()
                client_id = row[1]
                weight = row[2]
                price = row[3]

                # Если client_id — число с плавающей точкой, убираем ".0"
                if isinstance(client_id, float):
                    client_id = str(int(client_id))
                else:
                    client_id = str(client_id).strip()

                # Найти клиента по code
                client = Client.objects.filter(code=client_id).first()
                if not client:
                    messages.warning(request, f"Нету такого клиента: {client_id}")
                    continue

                # Получить статус "Бишкек"
                bishkek_status, _ = Status.objects.get_or_create(name=BaseStatus.BISHKEK)


                # Найти или обновить товар
                product, created = Product.objects.get_or_create(
                    product_code=product_code,
                    client=client,
                    defaults={
                        "date": datetime.utcnow(),
                        "status": bishkek_status,
                        "weight": weight,
                    }
                )

                if created:
                    products_created += 1
                else:
                    # Обновляем только эти значения
                    product.weight = weight
                    product.status = bishkek_status
                    product.price = price
                    product.date = datetime.utcnow()
                    product.save()
                    products_updated += 1

                # Увеличиваем счетчик товаров у клиента
                if client.id not in clients_products_count:
                    clients_products_count[client.id] = 0
                clients_products_count[client.id] += 1

            if products_created > 0:
                messages.success(request, f"Успешно создано товаров: {products_created}")
            if products_updated > 0:
                messages.info(request, f"Обновлено товаров: {products_updated}")
            
            # Добавляем сообщение с количеством товаров по каждому клиенту
            if clients_products_count:
                send_notification_bihskek.delay(clients_products_count)
                for client_id, count in clients_products_count.items():
                    messages.info(request, f"Клиент {client_id}: {count} товаров")

        except Exception as e:
            messages.error(request, f"Ошибка при обработке файла: {str(e)}")

        return redirect("page_bishkek")  # Перенаправляем на форму после обработки

    return redirect("page_bishkek")


@login_required
def add_product(request):
    # Получаем все статусы для формы
    statuses = Status.objects.all()

    if request.method == 'POST':
        # Получаем данные из формы
        product_code = request.POST.get('product_code')
        weight = request.POST.get('weight')
        price = request.POST.get('price')
        client_code = request.POST.get('client_code')
        status_id = request.POST.get('status')  # Получаем ID статуса из формы
        
        # Проверяем обязательные поля
        if not product_code or not client_code or not status_id:
            messages.error(request, "Код товара, код клиента и статус обязательны")
            return render(request, 'products/add_product.html', {
                'statuses': statuses,
                'current_page': 'add_product',
                'title': 'Добавить товар',
            })

        # Находим клиента
        try:
            client = Client.objects.get(code=client_code)
        except Client.DoesNotExist:
            messages.error(request, "Клиент с таким кодом не найден")
            return render(request, 'products/add_product.html', {
                'statuses': statuses,
                'current_page': 'add_product',
                'title': 'Добавить товар',
            })

        # Находим статус
        try:
            status = Status.objects.get(id=status_id)
        except Status.DoesNotExist:
            messages.error(request, "Выбранный статус не существует")
            return render(request, 'products/add_product.html', {
                'statuses': statuses,
                'current_page': 'add_product',
                'title': 'Добавить товар',
            })

        # Получаем филиал пользователя
        user = request.user
        branch = get_user_branch(user, Product).first().branch if get_user_branch(user, Product).exists() else None

        # Создаем товар
        product = Product(
            product_code=product_code,
            weight=weight if weight else None,
            price=price if price else None,
            date=timezone.now().date(),
            registered_at=timezone.now(),
            status=status,  # Используем выбранный статус
            client=client,
            branch=branch
        )
        product.save()

        # Перенаправляем на страницу добавления с сообщением об успехе
        messages.success(request, "Ваши изменения успешно сохранены!")
        return redirect('products')

    # Для GET-запроса рендерим форму с доступными статусами
    return render(request, 'products/add_product.html', {
        'statuses': statuses,
        'current_page': 'add_product',
        'title': 'Добавить товар',
    })